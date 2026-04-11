#!/usr/bin/env python3
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Iterable

import pymysql
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
TAXONOMY_XLSX = ROOT / "data" / "unclean" / "上中下游分类（一二三级分类标签）.xlsx"
TAGGING_XLSX = ROOT / "data" / "unclean" / "打标标签拆分结果（完整）.xlsx"
REPORT_PATH = ROOT / "SQL" / "tmp" / "industry-taxonomy-refresh-report-2026-04-10.md"

INDUSTRY_DIMENSION_NAME = "行业标签"
CHAIN_SUBDIMENSION_NAME = "产业链"
CATEGORY_SUBDIMENSION_NAME = "行业分类"

STAGE_META = {
    "上游": ("upstream", "上游 - 研发与技术"),
    "中游": ("midstream", "中游 - 产品与制造"),
    "下游": ("downstream", "下游 - 应用与服务"),
}
FALSE_VALUES = {"", "0", "0.0", "false", "False", "FALSE", "否", "无", "nan", "None"}


@dataclass(frozen=True)
class TaxonomyRow:
    stage_text: str
    level0: str
    level1: str
    level2: str
    level3: str


def normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\r", "").strip()


def truthy_cell(value) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, Decimal)):
        return float(value) != 0.0
    return normalize_text(value) not in FALSE_VALUES


def load_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    if not path.exists():
        return env
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip()
    return env


def connect_mysql() -> pymysql.connections.Connection:
    env = load_env(ROOT / ".env")
    return pymysql.connect(
        host=env.get("DB_HOST", "127.0.0.1"),
        user=env.get("DB_USER", "root"),
        password=env.get("DB_PASSWORD", ""),
        database=env.get("DB_NAME", "industrial_chain"),
        charset="utf8mb4",
        autocommit=False,
        cursorclass=pymysql.cursors.DictCursor,
    )


def read_taxonomy_rows() -> list[TaxonomyRow]:
    wb = load_workbook(TAXONOMY_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    next(rows)
    result: list[TaxonomyRow] = []
    for raw in rows:
        if not any(raw):
            continue
        row = TaxonomyRow(
            stage_text=normalize_text(raw[0]),
            level0=normalize_text(raw[1]),
            level1=normalize_text(raw[2]),
            level2=normalize_text(raw[3]),
            level3=normalize_text(raw[4]),
        )
        if not all([row.stage_text, row.level0, row.level1, row.level2, row.level3]):
            raise ValueError(f"taxonomy row is incomplete: {row}")
        if row.stage_text not in STAGE_META:
            raise ValueError(f"unknown stage: {row.stage_text}")
        result.append(row)
    if not result:
        raise ValueError("taxonomy xlsx has no usable rows")
    return result


def build_taxonomy_payload(rows: list[TaxonomyRow]) -> dict:
    level0_order: dict[str, int] = {}
    level1_order: dict[tuple[str, str], int] = {}
    level2_order: dict[tuple[str, str, str], int] = {}
    level3_order: dict[tuple[str, str, str, str], int] = {}

    for row in rows:
        level0_order.setdefault(row.level0, len(level0_order) + 1)
        level1_order.setdefault((row.level0, row.level1), sum(1 for key in level1_order if key[0] == row.level0) + 1)
        level2_order.setdefault(
            (row.level0, row.level1, row.level2),
            sum(1 for key in level2_order if key[:2] == (row.level0, row.level1)) + 1,
        )
        level3_order.setdefault(
            (row.level0, row.level1, row.level2, row.level3),
            sum(1 for key in level3_order if key[:3] == (row.level0, row.level1, row.level2)) + 1,
        )

    category_rows: list[dict] = []
    level0_code_by_name: dict[str, str] = {}
    level1_code_by_path: dict[tuple[str, str], str] = {}
    level2_code_by_path: dict[tuple[str, str, str], str] = {}
    level3_code_by_name: dict[str, str] = {}
    level1_paths_by_chain: dict[str, list[tuple[str, str]]] = defaultdict(list)
    leaf_infos_by_name: dict[str, list[dict]] = defaultdict(list)

    def add_category(name: str, level: int, code: str, parent_code: str | None) -> None:
        category_rows.append(
            {
                "category_name": name,
                "category_level": level,
                "category_level_code": code,
                "category_level_code_parent": parent_code,
                "field_belong": 1,
                "sort_order": int(code),
            }
        )

    for level0, order in level0_order.items():
        code = f"{order:02d}"
        level0_code_by_name[level0] = code
        add_category(level0, 0, code, None)

    for (level0, level1), order in level1_order.items():
        code = f"{level0_code_by_name[level0]}{order:02d}"
        level1_code_by_path[(level0, level1)] = code
        add_category(level1, 1, code, level0_code_by_name[level0])

    for (level0, level1, level2), order in level2_order.items():
        parent_code = level1_code_by_path[(level0, level1)]
        code = f"{parent_code}{order:02d}"
        level2_code_by_path[(level0, level1, level2)] = code
        add_category(level2, 2, code, parent_code)

    for row in rows:
        chain_key = (row.level0, row.level1)
        if chain_key not in level1_paths_by_chain[row.level1]:
            level1_paths_by_chain[row.level1].append(chain_key)

    seen_level3: set[tuple[str, str, str, str]] = set()
    for row in rows:
        if row.level3 in {"-", "—", "--"}:
            continue
        level3_key = (row.level0, row.level1, row.level2, row.level3)
        if level3_key in seen_level3:
            continue
        seen_level3.add(level3_key)
        parent_code = level2_code_by_path[(row.level0, row.level1, row.level2)]
        code = f"{parent_code}{level3_order[level3_key]:02d}"
        add_category(row.level3, 3, code, parent_code)
        level3_code_by_name[row.level3] = code
        leaf_info = {
            "stage_text": row.stage_text,
            "level0": row.level0,
            "level1": row.level1,
            "level2": row.level2,
            "level3": row.level3,
            "category_level_code": code,
        }
        if leaf_info not in leaf_infos_by_name[row.level3]:
            leaf_infos_by_name[row.level3].append(leaf_info)

    chain_rows: list[dict] = []
    for index, (chain_name, paths) in enumerate(level1_paths_by_chain.items(), start=1):
        first_level0, _ = paths[0]
        first_stage = next(row.stage_text for row in rows if row.level0 == first_level0 and row.level1 == chain_name)
        stage_key, stage_title = STAGE_META[first_stage]
        chain_rows.append(
            {
                "chain_name": chain_name,
                "stage_key": stage_key,
                "stage_title": stage_title,
                "chain_des": f"{first_stage}产业链方向：{chain_name}",
                "sort_order": index,
            }
        )

    return {
        "category_rows": category_rows,
        "chain_rows": chain_rows,
        "level1_paths_by_chain": level1_paths_by_chain,
        "leaf_infos_by_name": leaf_infos_by_name,
    }


def read_tagging_payload() -> dict:
    wb = load_workbook(TAGGING_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    if len(header) < 3:
        raise ValueError("tagging xlsx header is incomplete")
    tag_names = [normalize_text(value) for value in header[2:] if normalize_text(value)]
    companies: list[dict] = []
    for raw in rows:
        if not any(raw):
            continue
        company_name = normalize_text(raw[0])
        if not company_name:
            continue
        tags = []
        for tag_name, value in zip(tag_names, raw[2:]):
            if truthy_cell(value):
                tags.append(tag_name)
        companies.append({"company_name": company_name, "tags": tags})
    if not companies:
        raise ValueError("tagging xlsx has no usable company rows")
    return {"tag_names": tag_names, "companies": companies}


def fetch_company_name_map(cursor) -> dict[str, str]:
    cursor.execute(
        """
        SELECT company_name, credit_code
        FROM company_basic
        WHERE company_name IS NOT NULL AND company_name <> ''
        """
    )
    name_to_credit_code: dict[str, str] = {}
    for row in cursor.fetchall():
        company_name = normalize_text(row["company_name"])
        credit_code = normalize_text(row["credit_code"])
        if company_name in name_to_credit_code and name_to_credit_code[company_name] != credit_code:
            raise ValueError(f"duplicate company name in db: {company_name}")
        name_to_credit_code[company_name] = credit_code
    return name_to_credit_code


def ensure_industry_tag_dimensions(cursor) -> tuple[int, int, int]:
    cursor.execute(
        "SELECT company_tag_dimension_id FROM company_tag_dimension WHERE company_tag_dimension_name = %s LIMIT 1",
        [INDUSTRY_DIMENSION_NAME],
    )
    row = cursor.fetchone()
    if row:
        dimension_id = int(row["company_tag_dimension_id"])
    else:
        cursor.execute(
            """
            INSERT INTO company_tag_dimension (
              company_tag_dimension_name,
              company_tag_dimension_color,
              company_tag_dimension_icon,
              company_tag_dimension_des,
              sort_order
            ) VALUES (%s, %s, %s, %s, %s)
            """,
            [INDUSTRY_DIMENSION_NAME, "#fa8c16", "TagsOutlined", "行业标签维度", 10],
        )
        dimension_id = int(cursor.lastrowid)

    subdimension_ids: dict[str, int] = {}
    for sort_order, name in enumerate([CHAIN_SUBDIMENSION_NAME, CATEGORY_SUBDIMENSION_NAME], start=1):
        cursor.execute(
            """
            SELECT company_tag_subdimension_id
            FROM company_tag_subdimension
            WHERE company_tag_dimension_id = %s AND company_tag_subdimension_name = %s
            LIMIT 1
            """,
            [dimension_id, name],
        )
        row = cursor.fetchone()
        if row:
            subdimension_ids[name] = int(row["company_tag_subdimension_id"])
            continue
        cursor.execute(
            """
            INSERT INTO company_tag_subdimension (
              company_tag_subdimension_name,
              company_tag_dimension_id,
              sort_order
            ) VALUES (%s, %s, %s)
            """,
            [name, dimension_id, sort_order],
        )
        subdimension_ids[name] = int(cursor.lastrowid)
    return dimension_id, subdimension_ids[CHAIN_SUBDIMENSION_NAME], subdimension_ids[CATEGORY_SUBDIMENSION_NAME]


def chunked(rows: list[tuple], size: int = 1000) -> Iterable[list[tuple]]:
    for index in range(0, len(rows), size):
        yield rows[index : index + size]


def insert_many(cursor, sql: str, rows: list[tuple], batch_size: int = 1000) -> None:
    for batch in chunked(rows, batch_size):
        cursor.executemany(sql, batch)


def refresh_industry_data() -> dict:
    taxonomy_rows = read_taxonomy_rows()
    taxonomy_payload = build_taxonomy_payload(taxonomy_rows)
    tagging_payload = read_tagging_payload()

    missing_tag_headers = [tag_name for tag_name in tagging_payload["tag_names"] if tag_name not in taxonomy_payload["leaf_infos_by_name"]]
    if missing_tag_headers:
        raise ValueError(f"tagging xlsx contains unknown tags: {missing_tag_headers[:20]}")
    ambiguous_tag_headers = [
        tag_name
        for tag_name in tagging_payload["tag_names"]
        if len(taxonomy_payload["leaf_infos_by_name"][tag_name]) != 1
    ]
    if ambiguous_tag_headers:
        raise ValueError(f"tagging xlsx contains ambiguous tags: {ambiguous_tag_headers[:20]}")

    with connect_mysql() as connection:
        with connection.cursor() as cursor:
            name_to_credit_code = fetch_company_name_map(cursor)
            dimension_id, chain_subdimension_id, category_subdimension_id = ensure_industry_tag_dimensions(cursor)

            category_tag_names = []
            category_tag_sort_order: dict[str, int] = {}
            for sort_order, tag_name in enumerate(tagging_payload["tag_names"], start=1):
                if tag_name not in category_tag_sort_order:
                    category_tag_sort_order[tag_name] = sort_order
                    category_tag_names.append(tag_name)

            company_category_map_rows: list[tuple[int, str]] = []
            company_tag_map_rows: list[tuple[str, str, int, Decimal]] = []
            missing_companies: list[str] = []
            unmatched_tagged_company_count = 0
            tagged_company_count = 0
            total_leaf_assignments = 0
            total_chain_assignments = 0

            for company in tagging_payload["companies"]:
                credit_code = name_to_credit_code.get(company["company_name"])
                if not credit_code:
                    missing_companies.append(company["company_name"])
                    if company["tags"]:
                        unmatched_tagged_company_count += 1
                    continue
                if company["tags"]:
                    tagged_company_count += 1
                chain_tags_for_company: set[str] = set()
                for tag_name in company["tags"]:
                    leaf_info = taxonomy_payload["leaf_infos_by_name"][tag_name][0]
                    company_category_map_rows.append((0, credit_code, tag_name))
                    company_tag_map_rows.append((credit_code, tag_name, category_subdimension_id, Decimal("1.00")))
                    chain_tags_for_company.add(leaf_info["level1"])
                    total_leaf_assignments += 1
                for chain_tag in sorted(chain_tags_for_company):
                    company_tag_map_rows.append((credit_code, chain_tag, chain_subdimension_id, Decimal("1.00")))
                    total_chain_assignments += 1

            cursor.execute("SET FOREIGN_KEY_CHECKS = 0")
            cursor.execute("DELETE FROM company_tag_map WHERE company_tag_id IN (SELECT company_tag_id FROM company_tag_library WHERE company_tag_subdimension_id IN (%s, %s))", [chain_subdimension_id, category_subdimension_id])
            cursor.execute("DELETE FROM company_tag_library WHERE company_tag_subdimension_id IN (%s, %s)", [chain_subdimension_id, category_subdimension_id])
            cursor.execute("DELETE FROM chain_industry_category_industry_map")
            cursor.execute("DELETE FROM category_industry_company_map")
            cursor.execute("DELETE FROM chain_industry")
            cursor.execute("DELETE FROM category_industry")
            cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
            cursor.execute("SELECT company_tag_name FROM company_tag_library")
            reserved_tag_names = {normalize_text(row["company_tag_name"]) for row in cursor.fetchall()}
            insertable_chain_tag_rows = [
                row for row in taxonomy_payload["chain_rows"] if row["chain_name"] not in reserved_tag_names
            ]
            insertable_category_tag_names = [
                tag_name for tag_name in category_tag_names if tag_name not in reserved_tag_names
            ]
            skipped_tag_names = sorted(
                (
                    {row["chain_name"] for row in taxonomy_payload["chain_rows"]} | set(category_tag_names)
                )
                - ({row["chain_name"] for row in insertable_chain_tag_rows} | set(insertable_category_tag_names))
            )

            insert_many(
                cursor,
                """
                INSERT INTO category_industry (
                  category_name,
                  category_level,
                  category_level_code,
                  category_level_code_parent,
                  field_belong,
                  sort_order
                ) VALUES (%s, %s, %s, %s, %s, %s)
                """,
                [
                    (
                        row["category_name"],
                        row["category_level"],
                        row["category_level_code"],
                        row["category_level_code_parent"],
                        row["field_belong"],
                        row["sort_order"],
                    )
                    for row in taxonomy_payload["category_rows"]
                ],
            )
            cursor.execute("SELECT category_id, category_level_code FROM category_industry")
            category_id_by_code = {normalize_text(row["category_level_code"]): int(row["category_id"]) for row in cursor.fetchall()}

            insert_many(
                cursor,
                """
                INSERT INTO chain_industry (
                  chain_name,
                  stage_key,
                  stage_title,
                  chain_des,
                  sort_order
                ) VALUES (%s, %s, %s, %s, %s)
                """,
                [
                    (
                        row["chain_name"],
                        row["stage_key"],
                        row["stage_title"],
                        row["chain_des"],
                        row["sort_order"],
                    )
                    for row in taxonomy_payload["chain_rows"]
                ],
            )
            cursor.execute("SELECT chain_id, chain_name FROM chain_industry")
            chain_id_by_name = {normalize_text(row["chain_name"]): int(row["chain_id"]) for row in cursor.fetchall()}

            chain_category_map_rows = []
            for chain_name, level1_paths in taxonomy_payload["level1_paths_by_chain"].items():
                for level0, level1 in level1_paths:
                    level1_code = next(
                        row["category_level_code"]
                        for row in taxonomy_payload["category_rows"]
                        if row["category_level"] == 1
                        and row["category_name"] == level1
                        and row["category_level_code_parent"]
                        == next(
                            item["category_level_code"]
                            for item in taxonomy_payload["category_rows"]
                            if item["category_level"] == 0 and item["category_name"] == level0
                        )
                    )
                    chain_category_map_rows.append((chain_id_by_name[chain_name], category_id_by_code[level1_code]))
            insert_many(
                cursor,
                """
                INSERT INTO chain_industry_category_industry_map (chain_id, category_id)
                VALUES (%s, %s)
                """,
                chain_category_map_rows,
            )

            insert_many(
                cursor,
                """
                INSERT INTO category_industry_company_map (category_id, credit_code)
                VALUES (%s, %s)
                """,
                [
                    (category_id_by_code[taxonomy_payload["leaf_infos_by_name"][tag_name][0]["category_level_code"]], credit_code)
                    for _, credit_code, tag_name in company_category_map_rows
                ],
            )

            insert_many(
                cursor,
                """
                INSERT INTO company_tag_library (
                  company_tag_name,
                  company_tag_subdimension_id,
                  company_tag_level,
                  sort_order
                ) VALUES (%s, %s, NULL, %s)
                """,
                [
                    (row["chain_name"], chain_subdimension_id, row["sort_order"])
                    for row in insertable_chain_tag_rows
                ]
                + [
                    (tag_name, category_subdimension_id, category_tag_sort_order[tag_name])
                    for tag_name in insertable_category_tag_names
                ],
            )

            cursor.execute(
                """
                SELECT company_tag_id, company_tag_name, company_tag_subdimension_id
                FROM company_tag_library
                WHERE company_tag_subdimension_id IN (%s, %s)
                """,
                [chain_subdimension_id, category_subdimension_id],
            )
            tag_id_by_identity = {
                (normalize_text(row["company_tag_name"]), int(row["company_tag_subdimension_id"])): int(row["company_tag_id"])
                for row in cursor.fetchall()
            }
            insert_many(
                cursor,
                """
                INSERT INTO company_tag_map (credit_code, company_tag_id, source, confidence, user_id)
                VALUES (%s, %s, 2, %s, NULL)
                ON DUPLICATE KEY UPDATE source = VALUES(source), confidence = VALUES(confidence), create_time = CURRENT_TIMESTAMP
                """,
                [
                    (credit_code, tag_id_by_identity[(tag_name, subdimension_id)], confidence)
                    for credit_code, tag_name, subdimension_id, confidence in company_tag_map_rows
                    if (tag_name, subdimension_id) in tag_id_by_identity
                ],
            )
            connection.commit()

    report = {
        "taxonomy_row_count": len(taxonomy_rows),
        "category_count": len(taxonomy_payload["category_rows"]),
        "chain_count": len(taxonomy_payload["chain_rows"]),
        "chain_category_map_count": sum(len(paths) for paths in taxonomy_payload["level1_paths_by_chain"].values()),
        "category_tag_count": len(category_tag_names),
        "inserted_chain_tag_count": len(insertable_chain_tag_rows),
        "inserted_category_tag_count": len(insertable_category_tag_names),
        "skipped_conflicting_tag_count": len(skipped_tag_names),
        "tagging_company_count": len(tagging_payload["companies"]),
        "tagged_company_count": tagged_company_count,
        "missing_company_count": len(missing_companies),
        "unmatched_tagged_company_count": unmatched_tagged_company_count,
        "leaf_assignment_count": total_leaf_assignments,
        "chain_assignment_count": total_chain_assignments,
    }
    write_report(report, missing_companies, skipped_tag_names)
    return report


def write_report(report: dict, missing_companies: list[str], skipped_tag_names: list[str]) -> None:
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# Industry Taxonomy Refresh Report",
        "",
        f"- taxonomy rows: {report['taxonomy_row_count']}",
        f"- category_industry rows: {report['category_count']}",
        f"- chain_industry rows: {report['chain_count']}",
        f"- chain_industry_category_industry_map rows: {report['chain_category_map_count']}",
        f"- industry category tag rows: {report['category_tag_count']}",
        f"- inserted chain tags: {report['inserted_chain_tag_count']}",
        f"- inserted category tags: {report['inserted_category_tag_count']}",
        f"- skipped conflicting tags: {report['skipped_conflicting_tag_count']}",
        f"- tagging workbook companies: {report['tagging_company_count']}",
        f"- tagged companies: {report['tagged_company_count']}",
        f"- missing companies in db: {report['missing_company_count']}",
        f"- tagged companies missing in db: {report['unmatched_tagged_company_count']}",
        f"- leaf tag assignments: {report['leaf_assignment_count']}",
        f"- chain tag assignments: {report['chain_assignment_count']}",
        "",
    ]
    if missing_companies:
        lines.append("## Missing Companies")
        lines.append("")
        for company_name in missing_companies[:100]:
            lines.append(f"- {company_name}")
    if skipped_tag_names:
        lines.append("")
        lines.append("## Skipped Conflicting Tags")
        lines.append("")
        for tag_name in skipped_tag_names[:200]:
            lines.append(f"- {tag_name}")
    REPORT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    report = refresh_industry_data()
    print("Refreshed industry taxonomy and tag data.")
    for key, value in report.items():
        print(f"{key}: {value}")
    print(f"report_path: {REPORT_PATH}")


if __name__ == "__main__":
    main()
