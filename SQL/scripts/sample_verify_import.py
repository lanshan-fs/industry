#!/usr/bin/env python3
from __future__ import annotations

import argparse
import random
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import load_workbook

from _import_utils import normalize_company_name, parse_env, query_rows


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_XLSX = PROJECT_ROOT / "data" / "unclean" / "4.3全部企业数据汇总.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "SQL" / "reports" / f"import-sample-verification-{date.today().isoformat()}.md"


def load_workbook_aggregates(xlsx_path: Path) -> Dict[str, Dict[str, Any]]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    data: Dict[str, Dict[str, Any]] = defaultdict(
        lambda: {
            "basic": None,
            "operation": None,
            "ip": None,
            "risk": None,
            "subdistrict": [],
            "branches": [],
            "licenses": [],
            "qualifications": [],
            "customers": [],
            "rankings": [],
            "recruits": [],
            "software": [],
            "works": [],
            "patents": [],
        }
    )

    def rows(sheet_name: str) -> Iterable[Tuple[Any, ...]]:
        worksheet = workbook[sheet_name]
        iterator = worksheet.iter_rows(values_only=True)
        headers = list(next(iterator))
        for row in iterator:
            yield tuple(row[headers.index(col)] if col in headers else None for col in headers)

    sheet_headers: Dict[str, List[str]] = {}
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        sheet_headers[sheet_name] = list(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)))

    def as_dict(sheet_name: str, row: Tuple[Any, ...]) -> Dict[str, Any]:
        return dict(zip(sheet_headers[sheet_name], row))

    for raw_row in rows("企业基本信息"):
        row = as_dict("企业基本信息", raw_row)
        company_name = normalize_company_name(row.get("企业名称"))
        if company_name:
            data[company_name]["basic"] = row

    for raw_row in rows("企业经营信息"):
        row = as_dict("企业经营信息", raw_row)
        company_name = normalize_company_name(row.get("企业名称"))
        if company_name:
            data[company_name]["operation"] = row

    for raw_row in rows("知识产权"):
        row = as_dict("知识产权", raw_row)
        company_name = normalize_company_name(row.get("企业名称"))
        if company_name:
            data[company_name]["ip"] = row

    for raw_row in rows("扣分项"):
        row = as_dict("扣分项", raw_row)
        company_name = normalize_company_name(row.get("企业名称"))
        if company_name:
            data[company_name]["risk"] = row

    multi_specs = [
        ("分支机构", "企业名称", "branches"),
        ("许可", "企业名称", "licenses"),
        ("资质", "企业名称", "qualifications"),
        ("客户信息", "公司名称", "customers"),
        ("上榜榜单信息", "企业名称", "rankings"),
        ("招聘信息", "公司名称", "recruits"),
        ("软件著作权", "著作权人", "software"),
        ("作品著作权", "企业名称", "works"),
        ("专利信息", "申请人", "patents"),
        ("街道信息", "企业名称", "subdistrict"),
    ]
    for sheet_name, company_col, bucket in multi_specs:
        for raw_row in rows(sheet_name):
            row = as_dict(sheet_name, raw_row)
            company_name = normalize_company_name(row.get(company_col))
            if company_name:
                data[company_name][bucket].append(row)

    workbook.close()
    return data


def fetch_db_rows(env: Dict[str, str], company_names: List[str]) -> Dict[str, Dict[str, Any]]:
    escaped_names = ", ".join("'" + name.replace("\\", "\\\\").replace("'", "''") + "'" for name in company_names)
    base_rows = query_rows(
        f"""
        SELECT
          company_name,
          credit_code,
          legal_representative,
          register_number,
          org_code,
          DATE_FORMAT(establish_date, '%Y-%m-%d'),
          register_capital,
          paid_capital,
          company_type,
          org_type,
          company_scale,
          register_address,
          industry_belong,
          business_scope,
          email_business,
          taxpayer_credit_rating,
          COALESCE(branch_count, 0),
          COALESCE(recruit_count, 0),
          COALESCE(software_copyright_count, 0),
          COALESCE(work_copyright_count, 0),
          COALESCE(patent_count, 0),
          COALESCE(customer_count, 0),
          COALESCE(ranking_count, 0),
          COALESCE(supplier_count_raw, 0),
          COALESCE(qualification_count, 0),
          COALESCE(qualification_count_raw, 0),
          COALESCE(license_count, 0),
          COALESCE(license_count_raw, 0),
          COALESCE(legal_doc_all_count, 0),
          COALESCE(dishonest_execution_count, 0),
          COALESCE(business_abnormal_count, 0),
          COALESCE(bankruptcy_overlap_count, 0),
          COALESCE(executed_person_count, 0),
          COALESCE(consumption_restriction_count, 0),
          qualification_label
        FROM company_basic b
        JOIN company_basic_count c ON b.company_id = c.company_id
        WHERE company_name IN ({escaped_names})
        ORDER BY company_name;
        """,
        env,
    )
    company_ids = query_rows(f"SELECT company_name, company_id FROM company_basic WHERE company_name IN ({escaped_names});", env)
    company_id_by_name = {normalize_company_name(name): int(company_id) for name, company_id in company_ids}

    data: Dict[str, Dict[str, Any]] = {}
    for row in base_rows:
        name = normalize_company_name(row[0])
        if not name:
            continue
        data[name] = {
            "base": row,
            "company_id": company_id_by_name.get(name),
        }

    detail_queries = {
        "branches": "SELECT company_id, company_branch_name FROM company_branch WHERE company_id IN ({ids})",
        "licenses": "SELECT company_id, qualification_name FROM company_qualification WHERE record_kind = 'license' AND company_id IN ({ids})",
        "qualifications": "SELECT company_id, qualification_name FROM company_qualification WHERE record_kind = 'qualification' AND company_id IN ({ids})",
        "customers": "SELECT company_id, company_customer_name FROM company_customer WHERE company_id IN ({ids})",
        "rankings": "SELECT company_id, company_ranking_name FROM company_ranking WHERE company_id IN ({ids})",
        "recruits": "SELECT company_id, company_recruit_position FROM company_recruit WHERE company_id IN ({ids})",
        "software": "SELECT company_id, company_software_copyright_register_number FROM company_software_copyright WHERE company_id IN ({ids})",
        "works": "SELECT company_id, company_work_copyright_register_number FROM company_work_copyright WHERE company_id IN ({ids})",
        "patents": "SELECT company_id, company_patent_number FROM company_patent WHERE company_id IN ({ids})",
        "risks": "SELECT company_id, company_risk_category_name, company_risk_category_count FROM company_risk WHERE company_id IN ({ids})",
        "subdistrict": "SELECT company_id, company_subdistrict_name FROM company_subdistrict WHERE company_id IN ({ids})",
    }
    if not company_id_by_name:
        return data
    ids = ", ".join(str(company_id) for company_id in sorted(company_id_by_name.values()))
    names_by_id = {company_id: name for name, company_id in company_id_by_name.items()}

    for bucket, sql_template in detail_queries.items():
        rows = query_rows(sql_template.format(ids=ids), env)
        for row in rows:
            company_id = int(row[0])
            company_name = names_by_id.get(company_id)
            if not company_name:
                continue
            company_entry = data.setdefault(company_name, {"company_id": company_id})
            company_entry.setdefault(bucket, []).append(row[1:] if len(row) > 2 else row[1])
    return data


def choose_sample_companies(workbook_data: Dict[str, Dict[str, Any]], size: int, seed: int) -> List[str]:
    rich: List[str] = []
    basic_only: List[str] = []
    for company_name, entry in workbook_data.items():
        richness = sum(
            1
            for key in ["branches", "licenses", "qualifications", "customers", "rankings", "recruits", "software", "works", "patents"]
            if entry[key]
        )
        if richness >= 4:
            rich.append(company_name)
        else:
            basic_only.append(company_name)
    rng = random.Random(seed)
    rng.shuffle(rich)
    rng.shuffle(basic_only)
    rich_take = min(max(size // 2, 3), len(rich))
    sample = rich[:rich_take]
    remaining = size - len(sample)
    sample.extend(basic_only[:remaining])
    if len(sample) < size:
        pool = [name for name in rich[rich_take:] + basic_only[remaining:] if name not in sample]
        sample.extend(pool[: size - len(sample)])
    return sorted(sample)


def render_company_section(company_name: str, workbook_entry: Dict[str, Any], db_entry: Dict[str, Any]) -> List[str]:
    basic = workbook_entry.get("basic") or {}
    operation = workbook_entry.get("operation") or {}
    ip = workbook_entry.get("ip") or {}
    risk = workbook_entry.get("risk") or {}
    base = db_entry.get("base") or []
    db_risk_rows = db_entry.get("risks", [])
    db_risk_map = {row[0]: int(row[1]) for row in db_risk_rows}

    lines = [f"### {company_name}", ""]
    if not base:
        lines.extend(["- 结果: MySQL 未找到该企业", ""])
        return lines

    lines.extend(
        [
            "| 项目 | Excel | MySQL | 结论 |",
            "| --- | --- | --- | --- |",
            f"| 统一社会信用代码 | {basic.get('统一社会信用代码') or ''} | {base[1] or ''} | {'一致' if str(basic.get('统一社会信用代码') or '') == str(base[1] or '') else '不一致'} |",
            f"| 法定代表人 | {basic.get('法定代表人') or ''} | {base[2] or ''} | {'一致' if str(basic.get('法定代表人') or '') == str(base[2] or '') else '不一致'} |",
            f"| 注册号 | {basic.get('注册号') or ''} | {base[3] or ''} | {'一致' if str(basic.get('注册号') or '') == str(base[3] or '') else '不一致'} |",
            f"| 组织机构代码 | {basic.get('组织机构代码') or ''} | {base[4] or ''} | {'一致' if str(basic.get('组织机构代码') or '') == str(base[4] or '') else '不一致'} |",
            f"| 注册资本（万元） | {basic.get('注册资本（万元）') or ''} | {base[6] or ''} | {'近似一致' if str(basic.get('注册资本（万元）') or '').strip() and str(base[6] or '').strip() else '缺值'} |",
            f"| 实缴资本（万元） | {basic.get('实缴资本（万元）') or ''} | {base[7] or ''} | {'近似一致' if str(basic.get('实缴资本（万元）') or '').strip() and str(base[7] or '').strip() else '缺值'} |",
            f"| 分支机构数量 | {len(workbook_entry['branches'])} | {base[16]} | {'一致' if len(workbook_entry['branches']) == int(base[16]) else '不一致'} |",
            f"| 招聘数量 | {len(workbook_entry['recruits'])} | {base[17]} | {'一致' if len(workbook_entry['recruits']) == int(base[17]) else '不一致'} |",
            f"| 软件著作权数量 | {len(workbook_entry['software'])} | {base[18]} | {'一致' if len(workbook_entry['software']) == int(base[18]) else '不一致'} |",
            f"| 作品著作权数量 | {len(workbook_entry['works'])} | {base[19]} | {'一致' if len(workbook_entry['works']) == int(base[19]) else '不一致'} |",
            f"| 专利数量 | {len(workbook_entry['patents'])} | {base[20]} | {'一致' if len(workbook_entry['patents']) == int(base[20]) else '不一致'} |",
            f"| 客户数量 | {len(workbook_entry['customers'])} | {base[21]} | {'一致' if len(workbook_entry['customers']) == int(base[21]) else '不一致'} |",
            f"| 榜单数量 | {len(workbook_entry['rankings'])} | {base[22]} | {'一致' if len(workbook_entry['rankings']) == int(base[22]) else '不一致'} |",
            f"| 供应商数量（原始） | {operation.get('供应商数量') or 0} | {base[23]} | {'一致' if int(operation.get('供应商数量') or 0) == int(base[23]) else '不一致'} |",
            f"| 资质数量 | {len(workbook_entry['qualifications'])} / 原始 {operation.get('资质数量') or 0} | {base[24]} / 原始 {base[25]} | {'一致' if len(workbook_entry['qualifications']) == int(base[24]) and int(operation.get('资质数量') or 0) == int(base[25]) else '不一致'} |",
            f"| 许可数量 | {len(workbook_entry['licenses'])} / 原始 {operation.get('许可数量') or 0} | {base[26]} / 原始 {base[27]} | {'一致' if len(workbook_entry['licenses']) == int(base[26]) and int(operation.get('许可数量') or 0) == int(base[27]) else '不一致'} |",
            f"| 法律文书/司法案件 | {risk.get('司法案件') or 0} | {base[28]} | {'一致' if int(risk.get('司法案件') or 0) == int(base[28]) else '不一致'} |",
            f"| 失信被执行人 | {risk.get('失信被执行人') or 0} | {base[29]} | {'一致' if int(risk.get('失信被执行人') or 0) == int(base[29]) else '不一致'} |",
            f"| 经营异常 | {risk.get('经营异常') or 0} | {base[30]} | {'一致' if int(risk.get('经营异常') or 0) == int(base[30]) else '不一致'} |",
            f"| 破产案件 | {risk.get('破产案件') or 0} | {base[31]} | {'一致' if int(risk.get('破产案件') or 0) == int(base[31]) else '不一致'} |",
            f"| 被执行人 | {risk.get('被执行人') or 0} | {base[32]} | {'一致' if int(risk.get('被执行人') or 0) == int(base[32]) else '不一致'} |",
            f"| 限制高消费 | {risk.get('限制高消费') or 0} | {base[33]} | {'一致' if int(risk.get('限制高消费') or 0) == int(base[33]) else '不一致'} |",
        ]
    )

    lines.extend(
        [
            "",
            f"- Excel 分支样本: {', '.join(str(item.get('分支机构企业名称')) for item in workbook_entry['branches'][:3]) or '无'}",
            f"- MySQL 分支样本: {', '.join(str(item) for item in db_entry.get('branches', [])[:3]) or '无'}",
            f"- Excel 许可样本: {', '.join(str(item.get('许可文件名称')) for item in workbook_entry['licenses'][:3]) or '无'}",
            f"- MySQL 许可样本: {', '.join(str(item) for item in db_entry.get('licenses', [])[:3]) or '无'}",
            f"- Excel 资质样本: {', '.join(str(item.get('证书名称')) for item in workbook_entry['qualifications'][:3]) or '无'}",
            f"- MySQL 资质样本: {', '.join(str(item) for item in db_entry.get('qualifications', [])[:3]) or '无'}",
            f"- Excel 客户样本: {', '.join(str(item.get('客户名称')) for item in workbook_entry['customers'][:3]) or '无'}",
            f"- MySQL 客户样本: {', '.join(str(item) for item in db_entry.get('customers', [])[:3]) or '无'}",
            f"- Excel 榜单样本: {', '.join(str(item.get('榜单名称')) for item in workbook_entry['rankings'][:3]) or '无'}",
            f"- MySQL 榜单样本: {', '.join(str(item) for item in db_entry.get('rankings', [])[:3]) or '无'}",
            f"- Excel 招聘样本: {', '.join(str(item.get('招聘职位')) for item in workbook_entry['recruits'][:3]) or '无'}",
            f"- MySQL 招聘样本: {', '.join(str(item) for item in db_entry.get('recruits', [])[:3]) or '无'}",
            f"- Excel 软件著作权样本: {', '.join(str(item.get('登记号')) for item in workbook_entry['software'][:3]) or '无'}",
            f"- MySQL 软件著作权样本: {', '.join(str(item) for item in db_entry.get('software', [])[:3]) or '无'}",
            f"- Excel 作品著作权样本: {', '.join(str(item.get('登记号')) for item in workbook_entry['works'][:3]) or '无'}",
            f"- MySQL 作品著作权样本: {', '.join(str(item) for item in db_entry.get('works', [])[:3]) or '无'}",
            f"- Excel 专利样本: {', '.join(str(item.get('专利号')) for item in workbook_entry['patents'][:3]) or '无'}",
            f"- MySQL 专利样本: {', '.join(str(item) for item in db_entry.get('patents', [])[:3]) or '无'}",
            "",
        ]
    )
    return lines


def main() -> int:
    parser = argparse.ArgumentParser(description="Sample verify imported workbook data against MySQL.")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="Path to source workbook")
    parser.add_argument("--env-file", default=str(PROJECT_ROOT / ".env"), help="Path to .env file")
    parser.add_argument("--sample-size", type=int, default=10, help="Number of companies to sample")
    parser.add_argument("--seed", type=int, default=20260405, help="Random seed")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Markdown output path")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    env = parse_env(Path(args.env_file))
    workbook_data = load_workbook_aggregates(xlsx_path)
    sample_companies = choose_sample_companies(workbook_data, args.sample_size, args.seed)
    db_data = fetch_db_rows(env, sample_companies)

    lines = [
        "# 导入抽样核对报告",
        "",
        f"- 日期: {date.today().isoformat()}",
        f"- 样本数: {len(sample_companies)}",
        f"- 抽样种子: {args.seed}",
        "",
        "## 样本名单",
        "",
    ]
    lines.extend(f"- {name}" for name in sample_companies)
    lines.append("")
    lines.append("## 逐企业核对")
    lines.append("")

    for company_name in sample_companies:
        lines.extend(render_company_section(company_name, workbook_data[company_name], db_data.get(company_name, {})))

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"[ok] report written to {output_path}")
    print(f"[ok] sampled companies={len(sample_companies)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
