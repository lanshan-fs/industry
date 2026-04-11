#!/usr/bin/env python3
from __future__ import annotations

import argparse
import subprocess
import sys
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

from openpyxl import load_workbook

from _import_utils import insert_rows, parse_env, run_mysql_sql


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_UNCLEAN_DIR = ROOT / "data" / "unclean"


WorkbookConfig = Dict[str, object]


WORKBOOK_MAPPINGS: Dict[str, WorkbookConfig] = {
    "算法备案的医疗大模型(1).xlsx": {
        "sheets": {
            "Sheet1": {
                "table": "raw_import_company_ai_model_filing",
                "columns": [
                    ("时期", "period_raw"),
                    ("类型", "filing_type"),
                    ("序号", "source_order_raw"),
                    ("属地", "territory"),
                    ("名称", "model_name"),
                    ("单位", "company_name"),
                    ("编号", "filing_number"),
                    ("时间", "filed_at_raw"),
                ],
            }
        }
    },
    "高质量数据集.xlsx": {
        "sheets": {
            "Sheet1": {
                "table": "raw_import_company_high_quality_dataset",
                "columns": [
                    ("序号", "source_order_raw"),
                    ("案例名称", "dataset_name"),
                    ("申报单位", "applicant_unit"),
                    ("推荐单位", "recommender_unit"),
                ],
            }
        }
    },
    "专业能力评分-创新性.xlsx": {
        "sheets": {
            "创新医疗器械，突破性治疗公示": {
                "table": "raw_import_company_innovation_notice",
                "fixed": {"notice_type": "innovative_medical_device_beijing"},
                "columns": [
                    ("类别", "notice_category"),
                    ("原序号", "source_order_raw"),
                    ("时间", "public_date_raw"),
                    ("企业名称", "company_name"),
                    ("产品名称", "product_name"),
                    ("注册证号", "reg_no"),
                ],
            },
            "国家创新医疗器械公示": {
                "table": "raw_import_company_innovation_notice",
                "fixed": {"notice_type": "innovative_medical_device_national"},
                "columns": [
                    ("公示标题", "notice_title"),
                    ("产品名称", "product_name"),
                    ("申请人", "company_name"),
                ],
            },
            "药物拟纳入优先审评品种名单": {
                "table": "raw_import_company_innovation_notice",
                "fixed": {"notice_type": "priority_review_candidate"},
                "columns": [
                    ("序号", "source_order_raw"),
                    ("受理号", "acceptance_no"),
                    ("药品名称", "product_name"),
                    ("注册申请人", "company_name"),
                    ("公示日期", "public_date_raw"),
                    ("公示截止日期", "public_end_date_raw"),
                    ("是否为罕见病药物", "rare_disease_flag_raw"),
                ],
            },
            "突破性治疗公示": {
                "table": "raw_import_company_innovation_notice",
                "fixed": {"notice_type": "breakthrough_therapy"},
                "synthetic_headers": ["序号", "受理号", "药品名称", "注册申请人", "申请日期", "公示日期", "公示截止日期", "是否为罕见病药物"],
                "columns": [
                    ("序号", "source_order_raw"),
                    ("受理号", "acceptance_no"),
                    ("药品名称", "product_name"),
                    ("注册申请人", "company_name"),
                    ("公示日期", "public_date_raw"),
                    ("公示截止日期", "public_end_date_raw"),
                    ("是否为罕见病药物", "rare_disease_flag_raw"),
                ],
            },
            "主文档登记信息公示": {
                "table": "raw_import_company_innovation_notice",
                "fixed": {"notice_type": "master_file"},
                "columns": [
                    ("序号", "source_order_raw"),
                    ("所有者名称", "owner_name"),
                    ("主文档登记事项名称", "product_name"),
                    ("主文档登记号", "reg_no"),
                    ("登记/更新时间", "public_date_raw"),
                ],
            },
        }
    },
}


RAW_TABLES = [
    "raw_import_company_ai_model_filing",
    "raw_import_company_high_quality_dataset",
    "raw_import_company_innovation_notice",
]


def _load_rows(
    workbook_path: Path,
    sheet_name: str,
    columns: Sequence[Tuple[str, str]],
    synthetic_headers: Sequence[str] | None = None,
) -> List[Dict[str, object]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    iterator = worksheet.iter_rows(values_only=True)
    first_row = next(iterator, None)
    if first_row is None:
        workbook.close()
        return []

    if synthetic_headers:
        headers = list(synthetic_headers)
        normalized_first = [str(item).strip() if item is not None else "" for item in first_row]
        normalized_headers = [str(item).strip() for item in headers]
        if normalized_first == normalized_headers[: len(normalized_first)]:
            data_rows = list(iterator)
        else:
            data_rows = [first_row, *iterator]
    else:
        headers = list(first_row)
        data_rows = list(iterator)

    header_map = {header: idx for idx, header in enumerate(headers)}
    missing = [excel_name for excel_name, _ in columns if excel_name not in header_map]
    if missing:
        workbook.close()
        raise KeyError(f"{workbook_path.name}/{sheet_name} 缺少列: {missing}")

    rows: List[Dict[str, object]] = []
    for row_index, row in enumerate(data_rows, start=2 if not synthetic_headers else 1):
        if row is None:
            continue
        values = {db_name: row[header_map[excel_name]] for excel_name, db_name in columns}
        if all(value is None or str(value).strip() == "" for value in values.values()):
            continue
        values["sheet_row_no"] = row_index
        rows.append(values)
    workbook.close()
    return rows


def _truncate_tables(env: Dict[str, str]) -> None:
    sql = ["SET FOREIGN_KEY_CHECKS = 0;"]
    for table in RAW_TABLES:
        sql.append(f"TRUNCATE TABLE `{table}`;")
    sql.append("SET FOREIGN_KEY_CHECKS = 1;")
    run_mysql_sql("\n".join(sql) + "\n", env)


def import_workbook(workbook_path: Path, env: Dict[str, str], batch_size: int) -> Dict[str, int]:
    config = WORKBOOK_MAPPINGS[workbook_path.name]
    results: Dict[str, int] = {}
    for sheet_name, sheet_config in config["sheets"].items():
        rows = _load_rows(
            workbook_path,
            sheet_name,
            sheet_config["columns"],
            sheet_config.get("synthetic_headers"),
        )
        table = sheet_config["table"]
        fixed_values = sheet_config.get("fixed", {})

        insert_payload = []
        for row in rows:
            payload = {
                "source_file": workbook_path.name,
                "source_sheet": sheet_name,
                **fixed_values,
                **row,
            }
            insert_payload.append(payload)

        if insert_payload:
            columns = list(insert_payload[0].keys())
            insert_rows(
                env,
                table,
                columns,
                [tuple(item[column] for column in columns) for item in insert_payload],
                batch_size=batch_size,
            )
        results[f"{sheet_name}->{table}"] = len(insert_payload)
    return results


def main() -> int:
    parser = argparse.ArgumentParser(description="Import auxiliary Excel workbooks into raw_import_* tables.")
    parser.add_argument("--env-file", default=".env", help="Path to .env file")
    parser.add_argument("--unclean-dir", default=str(DEFAULT_UNCLEAN_DIR), help="Directory containing auxiliary xlsx files")
    parser.add_argument("--batch-size", type=int, default=500, help="INSERT batch size")
    parser.add_argument("--no-truncate", action="store_true", help="Do not truncate target raw_import_* tables before import")
    args = parser.parse_args()

    env_path = Path(args.env_file)
    if not env_path.exists():
        raise FileNotFoundError(f".env 文件不存在: {env_path}")
    env = parse_env(env_path)

    unclean_dir = Path(args.unclean_dir)
    missing = [name for name in WORKBOOK_MAPPINGS if not (unclean_dir / name).exists()]
    if missing:
        raise FileNotFoundError(f"未找到辅助 Excel: {missing}")

    if not args.no_truncate:
        print("[import-aux] truncating raw_import auxiliary tables")
        _truncate_tables(env)

    for workbook_name in WORKBOOK_MAPPINGS:
        workbook_path = unclean_dir / workbook_name
        result = import_workbook(workbook_path, env, args.batch_size)
        for target, count in result.items():
            print(f"[import-aux] {workbook_name} {target}: {count}")

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except subprocess.CalledProcessError as exc:
        print("[error] mysql command failed", file=sys.stderr)
        raise SystemExit(exc.returncode) from exc
