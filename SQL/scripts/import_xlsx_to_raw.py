#!/usr/bin/env python3
from __future__ import annotations

import argparse
import subprocess
import sys
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple

from openpyxl import load_workbook

from _import_utils import chunked, escape_sql, parse_env, run_mysql_sql


SHEET_MAPPINGS: Dict[str, Dict[str, object]] = {
    "企业基本信息": {
        "table": "raw_import_company_basic",
        "columns": [
            ("序号", "sheet_row_no"),
            ("企业名称", "company_name"),
            ("统一社会信用代码", "credit_code"),
            ("成立年限", "establish_info"),
            ("注册资本（万元）", "register_capital_raw"),
            ("实缴资本（万元）", "paid_capital_raw"),
            ("企业类型", "company_type_raw"),
            ("组织类型", "org_type_raw"),
            ("企业规模", "company_scale_raw"),
            ("是否有分支机构", "has_branch_raw"),
            ("分支机构数量", "branch_count_raw"),
            ("地址信息", "address_info_raw"),
            ("投融资轮次", "financing_round_raw"),
            ("法定代表人", "legal_representative"),
            ("社保人数", "insured_count_raw"),
            ("注册号", "register_number"),
            ("组织机构代码", "org_code"),
            ("所属行业", "industry_name_raw"),
            ("经营范围", "business_scope_raw"),
            ("邮箱（工商信息）", "email_business"),
            ("股东", "shareholder_raw"),
            ("联系电话", "contact_phone_0"),
            ("联系电话1", "contact_phone_1"),
            ("联系电话2", "contact_phone_2"),
            ("联系电话3", "contact_phone_3"),
            ("联系电话4", "contact_phone_4"),
            ("联系电话5", "contact_phone_5"),
        ],
    },
    "分支机构": {
        "table": "raw_import_company_branch",
        "columns": [
            ("企业名称", "company_name"),
            ("分支机构企业名称", "branch_company_name"),
            ("负责人", "branch_leader"),
            ("地区", "branch_region"),
            ("成立日期", "branch_establish_date_raw"),
            ("登记状态", "branch_status_raw"),
        ],
    },
    "企业经营信息": {
        "table": "raw_import_company_operation",
        "columns": [
            ("序号", "sheet_row_no"),
            ("企业名称", "company_name"),
            ("员工人数", "employee_count_raw"),
            ("社保人数", "insured_count_raw"),
            ("上市状态", "listing_status_raw"),
            ("国标行业", "national_industry_raw"),
            ("联系方式", "contact_info_raw"),
            ("同企业电话", "same_phone_count_raw"),
            ("邮箱（工商信息）_y", "email_business"),
            ("是否小微企业", "is_micro_enterprise_raw"),
            ("是否有变更信息", "changed_info_raw"),
            ("是否为一般纳税人", "is_general_taxpayer_raw"),
            ("有无融资信息", "has_financing_info_raw"),
            ("有无招投标", "has_bidding_raw"),
            ("招投标数量", "bidding_count_raw"),
            ("有无供应商", "has_supplier_raw"),
            ("供应商数量", "supplier_count_raw"),
            ("有无招聘", "has_recruitment_raw"),
            ("招聘信息数量", "recruit_count_raw"),
            ("是否有客户信息", "has_customer_info_raw"),
            ("客户数量", "customer_count_raw"),
            ("是否有上榜榜单", "has_ranking_raw"),
            ("上榜榜单数量", "ranking_count_raw"),
            ("是否有许可", "has_license_raw"),
            ("许可数量", "license_count_raw"),
            ("是否有资质", "has_qualification_raw"),
            ("资质数量", "qualification_count_raw"),
            ("税务评级", "taxpayer_credit_rating_raw"),
        ],
    },
    "客户信息": {
        "table": "raw_import_company_customer",
        "columns": [
            ("序号", "sheet_row_no"),
            ("公司名称", "company_name"),
            ("客户名称", "customer_name"),
            ("销售占比", "sales_ratio_raw"),
            ("销售金额", "sales_amount_raw"),
            ("报告期", "report_period_raw"),
            ("数据来源", "data_source"),
        ],
    },
    "上榜榜单信息": {
        "table": "raw_import_company_ranking",
        "columns": [
            ("企业名称", "company_name"),
            ("榜单名称", "ranking_name"),
            ("榜单类型", "ranking_type"),
            ("来源", "ranking_source"),
            ("榜内位置", "ranking_position_raw"),
            ("榜内名称", "ranking_alias"),
            ("发布年份", "publish_year_raw"),
        ],
    },
    "招聘信息": {
        "table": "raw_import_company_recruit",
        "columns": [
            ("公司名称", "company_name"),
            ("招聘职位", "position_name"),
            ("薪资", "salary_raw"),
            ("地区", "work_place"),
            ("工作经验", "work_year_raw"),
            ("学历", "edu_req_raw"),
            ("发布时间", "recruit_time_raw"),
        ],
    },
    "许可": {
        "table": "raw_import_company_license",
        "columns": [
            ("企业名称", "company_name"),
            ("许可文件编号", "license_number"),
            ("许可文件名称", "license_name"),
            ("许可状态", "license_status"),
            ("来源", "data_source"),
            ("有效期", "validity_period_raw"),
            ("许可机关", "issuing_authority"),
        ],
    },
    "资质": {
        "table": "raw_import_company_qualification",
        "columns": [
            ("企业名称", "company_name"),
            ("证书名称", "qualification_name"),
            ("证书编号", "qualification_number"),
            ("证书类型", "qualification_type"),
            ("证书状态", "qualification_status"),
            ("发证日期", "issued_at_raw"),
            ("截止日期", "expires_at_raw"),
        ],
    },
    "知识产权": {
        "table": "raw_import_company_ip_overview",
        "columns": [
            ("序号", "sheet_row_no"),
            ("企业名称", "company_name"),
            ("有无作品著作", "has_work_copyright_raw"),
            ("作品著作权数量", "work_copyright_count_raw"),
            ("有无软件著作", "has_software_copyright_raw"),
            ("软件著作权数量", "software_copyright_count_raw"),
            ("高新技术企业", "is_high_tech_enterprise_raw"),
            ("专精特新中小企业", "is_srdi_sme_raw"),
            ("瞪羚企业", "is_gazelle_company_raw"),
            ("科技型中小企业", "is_tech_sme_raw"),
            ("雏鹰企业", "is_egalet_company_raw"),
            ("专精特新小巨人", "is_srdi_little_giant_raw"),
            ("创新型中小企业", "is_innovative_sme_raw"),
            ("有无专利", "has_patent_raw"),
            ("专利数量", "patent_count_raw"),
        ],
    },
    "软件著作权": {
        "table": "raw_import_software_copyright",
        "columns": [
            ("序号", "sheet_row_no"),
            ("软件名称", "software_name"),
            ("登记号", "register_number"),
            ("软件简称", "software_short_name"),
            ("登记批准日期", "register_date_raw"),
            ("状态", "status_raw"),
            ("取得方式", "obtain_method_raw"),
            ("著作权人", "company_name"),
        ],
    },
    "作品著作权": {
        "table": "raw_import_company_work_copyright",
        "columns": [
            ("序号", "sheet_row_no"),
            ("作品名称", "work_name"),
            ("登记号", "register_number"),
            ("类别", "work_type_raw"),
            ("首次发布日期", "first_publish_date_raw"),
            ("登记日期", "register_date_raw"),
            ("状态", "status_raw"),
            ("企业名称", "company_name"),
        ],
    },
    "专利信息": {
        "table": "raw_import_company_patent",
        "columns": [
            ("序号", "sheet_row_no"),
            ("申请人", "company_name"),
            ("专利号", "patent_number"),
            ("专利名称", "patent_name"),
            ("专利类型", "patent_type_raw"),
            ("申请日", "application_date_raw"),
            ("申请公布日", "publication_date_raw"),
        ],
    },
    "扣分项": {
        "table": "raw_import_company_risk",
        "columns": [
            ("企业名称", "company_name"),
            ("严重违法", "serious_illegal_count_raw"),
            ("司法案件", "judicial_case_count_raw"),
            ("合作风险", "cooperation_risk_count_raw"),
            ("失信被执行人", "dishonest_execution_count_raw"),
            ("破产案件", "bankruptcy_case_count_raw"),
            ("被执行人", "executed_person_count_raw"),
            ("限制高消费", "consumption_restriction_count_raw"),
            ("经营异常", "business_abnormal_count_raw"),
            ("集群注册", "cluster_registration_count_raw"),
        ],
    },
    "街道信息": {
        "table": "raw_import_company_subdistrict",
        "columns": [
            ("序号", "sheet_row_no"),
            ("企业名称", "company_name"),
            ("地址", "address_raw"),
            ("街道", "street_name"),
            ("地区", "region_name"),
        ],
    },
}


def load_sheet_rows(xlsx_path: Path, sheet_name: str, columns: Sequence[Tuple[str, str]]) -> List[Tuple[object, ...]]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    headers = list(next(rows))
    header_map = {header: idx for idx, header in enumerate(headers)}
    missing = [excel_name for excel_name, _ in columns if excel_name not in header_map]
    if missing:
        workbook.close()
        raise KeyError(f"{sheet_name} 缺少列: {missing}")

    results: List[Tuple[object, ...]] = []
    for row in rows:
        values = tuple(row[header_map[excel_name]] for excel_name, _ in columns)
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        results.append(values)
    workbook.close()
    return results


def import_sheet(xlsx_path: Path, env: Dict[str, str], sheet_name: str, batch_size: int) -> int:
    config = SHEET_MAPPINGS[sheet_name]
    table = str(config["table"])
    columns = list(config["columns"])  # type: ignore[arg-type]
    rows = load_sheet_rows(xlsx_path, sheet_name, columns)
    column_sql = ", ".join(f"`{column_name}`" for _, column_name in columns)

    for batch in chunked(rows, batch_size):
        values_sql = []
        for row in batch:
            values_sql.append("(" + ", ".join(escape_sql(value) for value in row) + ")")
        sql = f"INSERT INTO `{table}` ({column_sql}) VALUES\n" + ",\n".join(values_sql) + ";\n"
        run_mysql_sql(sql, env)
    return len(rows)


def truncate_raw_tables(env: Dict[str, str]) -> None:
    statements = ["SET FOREIGN_KEY_CHECKS = 0;"]
    for config in SHEET_MAPPINGS.values():
        statements.append(f"TRUNCATE TABLE `{config['table']}`;")
    statements.append("SET FOREIGN_KEY_CHECKS = 1;")
    run_mysql_sql("\n".join(statements) + "\n", env)


def apply_init_sql(init_sql_path: Path, env: Dict[str, str]) -> None:
    run_mysql_sql(init_sql_path.read_text(encoding="utf-8"), env, with_database=False)


def ensure_raw_schema(raw_schema_sql_path: Path, env: Dict[str, str]) -> None:
    run_mysql_sql(raw_schema_sql_path.read_text(encoding="utf-8"), env)


def main() -> int:
    parser = argparse.ArgumentParser(description="Import workbook sheets into raw_import_* tables.")
    parser.add_argument("--xlsx", default="data/unclean/4.3全部企业数据汇总.xlsx", help="Path to source xlsx file")
    parser.add_argument("--env-file", default=".env", help="Path to .env file")
    parser.add_argument("--init-sql", default="SQL/sql/init.sql", help="Path to init.sql")
    parser.add_argument("--raw-schema-sql", default="SQL/sql/raw_import_staging.sql", help="Path to raw staging schema SQL")
    parser.add_argument("--apply-init", action="store_true", help="Apply init.sql before import")
    parser.add_argument("--no-truncate", action="store_true", help="Do not truncate raw_import_* tables before import")
    parser.add_argument("--batch-size", type=int, default=400, help="INSERT batch size")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    env_path = Path(args.env_file)
    init_sql_path = Path(args.init_sql)
    raw_schema_sql_path = Path(args.raw_schema_sql)

    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX 文件不存在: {xlsx_path}")
    if not env_path.exists():
        raise FileNotFoundError(f".env 文件不存在: {env_path}")

    env = parse_env(env_path)

    if args.apply_init:
        print(f"[schema] applying {init_sql_path}")
        apply_init_sql(init_sql_path, env)

    print(f"[schema] ensuring raw staging schema via {raw_schema_sql_path}")
    ensure_raw_schema(raw_schema_sql_path, env)

    if not args.no_truncate:
        print("[import] truncating raw_import_* tables")
        truncate_raw_tables(env)

    total = 0
    for sheet_name in SHEET_MAPPINGS:
        count = import_sheet(xlsx_path, env, sheet_name, args.batch_size)
        total += count
        print(f"[import] {sheet_name} -> {SHEET_MAPPINGS[sheet_name]['table']}: {count}")

    print(f"[done] total rows imported: {total}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except subprocess.CalledProcessError as exc:
        print("[error] mysql command failed", file=sys.stderr)
        raise SystemExit(exc.returncode) from exc
