from __future__ import annotations

import csv
import json
import os
import re
import subprocess
from collections import defaultdict
from copy import deepcopy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
SQL_DIR = ROOT / "SQL"
DATA_DIR = SQL_DIR / "data"
DESIGN_DIR = SQL_DIR / "mysql-design"
SQL_OUT_DIR = SQL_DIR / "sql"
REPORT_DIR = SQL_DIR / "reports"
UNCLEAN_XLSX = ROOT / "data" / "unclean" / "4.3全部企业数据汇总.xlsx"

DESIGN_V1_PATH = DESIGN_DIR / "design-2026-4-5-V1.csv"
DESIGN_V2_PATH = DESIGN_DIR / "design-2026-4-5-V2.csv"
INIT_V1_PATH = SQL_OUT_DIR / "init-2026-4-5-V1.sql"
INIT_V2_PATH = SQL_OUT_DIR / "init.sql"
REPORT_PATH = REPORT_DIR / "design-2026-4-5-V2-changes.md"

SEED_TABLES_V1 = [
    "category_industry",
    "chain_industry",
    "chain_industry_category_industry_map",
    "company_tag_dimension",
    "company_tag_subdimension",
    "company_tag_library",
    "company_tag_auto_rule",
]

SEED_TABLES_V2 = [
    "category_industry",
    "chain_industry",
    "chain_industry_category_industry_map",
    "company_tag_dimension",
    "company_tag_subdimension",
    "company_tag_library",
    "company_tag_auto_rule",
]

EXTRA_TAG_DIMENSIONS = [
    {
        "company_tag_dimension_id": 10,
        "company_tag_dimension_name": "行业标签",
        "company_tag_dimension_color": None,
        "company_tag_dimension_icon": None,
        "company_tag_dimension_des": "产业链与行业分类标签",
        "sort_order": 61,
    }
]

EXTRA_TAG_SUBDIMENSIONS = [
    {
        "company_tag_subdimension_id": 63,
        "company_tag_subdimension_name": "产业链",
        "company_tag_dimension_id": 10,
        "sort_order": 1,
    },
    {
        "company_tag_subdimension_id": 64,
        "company_tag_subdimension_name": "行业分类",
        "company_tag_dimension_id": 10,
        "sort_order": 2,
    },
]

EXTRA_TAG_LIBRARY = [
    (315, "智慧医疗", 63, None, 1),
    (316, "互联网+健康", 63, None, 2),
    (317, "数字疗法", 63, None, 3),
    (318, "前沿技术融合", 63, None, 4),
    (319, "AI 药物研发平台", 63, None, 5),
    (320, "AI CRO / 技术服务商", 63, None, 6),
    (321, "AI 自研管线企业", 63, None, 7),
    (322, "AI 软件 / 工具平台", 63, None, 8),
    (323, "AI + 特定领域研发", 63, None, 9),
    (324, "体外诊断 (IVD)", 63, None, 10),
    (325, "影像设备", 63, None, 11),
    (326, "治疗设备", 63, None, 12),
    (327, "生命信息支持设备", 63, None, 13),
    (328, "康复设备", 63, None, 14),
    (329, "辅助设备", 63, None, 15),
    (330, "家用医疗设备", 63, None, 16),
    (331, "高值医用耗材", 63, None, 17),
    (332, "植入器械/材料", 63, None, 18),
    (333, "低值医用耗材", 63, None, 19),
    (334, "装备制造", 63, None, 20),
    (335, "化学制药", 63, None, 21),
    (336, "生物制品", 63, None, 22),
    (337, "中药", 63, None, 23),
    (338, "医药商业 / 流通", 63, None, 24),
    (339, "医疗零售", 63, None, 25),
    (340, "严肃医疗", 63, None, 26),
    (341, "消费医疗", 63, None, 27),
    (342, "互联网医疗", 63, None, 28),
    (343, "第三方中心", 63, None, 29),
    (344, "保险支付", 63, None, 30),
    (345, "智慧医疗（0101）", 64, None, 2),
    (346, "前沿技术融合（0201）", 64, None, 127),
    (347, "体外诊断 (IVD)（0301）", 64, None, 138),
    (348, "医药商业 / 流通（0401）", 64, None, 252),
    (349, "化学制药（0501）", 64, None, 313),
    (350, "互联网+健康（0102）", 64, None, 46),
    (351, "影像设备（0302）", 64, None, 151),
    (352, "医疗零售（0402）", 64, None, 260),
    (353, "生物制品（0502）", 64, None, 326),
    (354, "数字疗法（0103）", 64, None, 88),
    (355, "治疗设备（0303）", 64, None, 166),
    (356, "严肃医疗（0403）", 64, None, 271),
    (357, "中药（0503）", 64, None, 353),
    (358, "消费医疗（0404）", 64, None, 279),
    (359, "康复设备（0305）", 64, None, 190),
    (360, "互联网医疗（0405）", 64, None, 291),
    (361, "第三方中心（0406）", 64, None, 298),
    (362, "保险支付（0407）", 64, None, 307),
    (363, "植入器械/材料（0309）", 64, None, 221),
    (364, "低值医用耗材（0310）", 64, None, 239),
    (365, "装备制造（0311）", 64, None, 246),
]


def load_env_file(path: Path) -> dict[str, str]:
    result: dict[str, str] = {}
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        result[key.strip()] = value.strip()
    return result


ENV = load_env_file(ROOT / ".env")


def mysql_exec(query: str) -> list[list[str]]:
    host = "127.0.0.1"
    port = ENV.get("DB_PORT", "3306").strip() or "3306"
    user = ENV["DB_USER"].strip()
    password = ENV["DB_PASSWORD"].strip()
    database = ENV["DB_NAME"].strip()
    cmd = [
        "mysql",
        "-h",
        host,
        "-P",
        port,
        "-u",
        user,
        f"-p{password}",
        "-D",
        database,
        "-N",
        "-B",
        "-e",
        query,
    ]
    env = os.environ.copy()
    env["MYSQL_PWD"] = password
    proc = subprocess.run(cmd, cwd=ROOT, capture_output=True, text=True, check=True, env=env)
    return [line.split("\t") for line in proc.stdout.splitlines() if line]


def decode_show_create(raw_sql: str) -> str:
    sql = raw_sql.replace("\\n", "\n").replace("\\t", "\t")
    sql = re.sub(r"AUTO_INCREMENT=\d+\s*", "", sql)
    return sql


def get_show_create(table_name: str) -> str:
    rows = mysql_exec(f"SHOW CREATE TABLE `{table_name}`;")
    return decode_show_create(rows[0][1])


def sql_literal(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, dict):
        value = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    value = str(value).replace("\\", "\\\\").replace("'", "\\'")
    return f"'{value}'"


def insert_statement(table_name: str, rows: list[dict[str, Any]]) -> str:
    if not rows:
        return ""
    columns = list(rows[0].keys())
    values_sql = []
    for row in rows:
        values_sql.append("(" + ", ".join(sql_literal(row.get(column)) for column in columns) + ")")
    cols_sql = ", ".join(f"`{column}`" for column in columns)
    return f"INSERT INTO `{table_name}` ({cols_sql}) VALUES\n  " + ",\n  ".join(values_sql) + ";\n"


def load_json(path: Path) -> list[dict[str, Any]]:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: list[dict[str, Any]]) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def sync_tag_source_files() -> dict[str, int]:
    dimensions = load_json(DATA_DIR / "company_tag_dimension.json")
    subdimensions = load_json(DATA_DIR / "company_tag_subdimension.json")
    library = load_json(DATA_DIR / "company_tag_library.json")

    dimension_ids = {row["company_tag_dimension_id"] for row in dimensions}
    subdimension_ids = {row["company_tag_subdimension_id"] for row in subdimensions}
    library_ids = {row["company_tag_id"] for row in library}

    for row in EXTRA_TAG_DIMENSIONS:
        if row["company_tag_dimension_id"] not in dimension_ids:
            dimensions.append(row)
    for row in EXTRA_TAG_SUBDIMENSIONS:
        if row["company_tag_subdimension_id"] not in subdimension_ids:
            subdimensions.append(row)
    for tag_id, tag_name, subdimension_id, tag_level, sort_order in EXTRA_TAG_LIBRARY:
        if tag_id not in library_ids:
            library.append(
                {
                    "company_tag_id": tag_id,
                    "company_tag_name": tag_name,
                    "company_tag_subdimension_id": subdimension_id,
                    "company_tag_level": tag_level,
                    "sort_order": sort_order,
                }
            )

    dimensions.sort(key=lambda item: item["company_tag_dimension_id"])
    subdimensions.sort(key=lambda item: item["company_tag_subdimension_id"])
    library.sort(key=lambda item: item["company_tag_id"])

    write_json(DATA_DIR / "company_tag_dimension.json", dimensions)
    write_json(DATA_DIR / "company_tag_subdimension.json", subdimensions)
    write_json(DATA_DIR / "company_tag_library.json", library)

    return {
        "dimensions": len(dimensions),
        "subdimensions": len(subdimensions),
        "library": len(library),
    }


def live_schema_metadata() -> dict[str, Any]:
    table_rows = mysql_exec(
        """
        SELECT table_name, table_comment
        FROM information_schema.tables
        WHERE table_schema = DATABASE()
        ORDER BY table_name;
        """
    )
    tables = [row[0] for row in table_rows]
    table_comments = {row[0]: row[1] for row in table_rows}

    column_rows = mysql_exec(
        """
        SELECT
          table_name,
          column_name,
          column_type,
          is_nullable,
          column_default,
          column_key,
          extra,
          column_comment,
          ordinal_position
        FROM information_schema.columns
        WHERE table_schema = DATABASE()
        ORDER BY table_name, ordinal_position;
        """
    )

    foreign_key_rows = mysql_exec(
        """
        SELECT
          table_name,
          column_name,
          referenced_table_name,
          referenced_column_name
        FROM information_schema.key_column_usage
        WHERE table_schema = DATABASE()
          AND referenced_table_name IS NOT NULL
        ORDER BY table_name, column_name;
        """
    )

    unique_rows = mysql_exec(
        """
        SELECT s.table_name, s.column_name
        FROM information_schema.statistics s
        JOIN (
          SELECT table_schema, table_name, index_name
          FROM information_schema.statistics
          WHERE table_schema = DATABASE()
            AND non_unique = 0
            AND index_name <> 'PRIMARY'
          GROUP BY table_schema, table_name, index_name
          HAVING COUNT(*) = 1
        ) u
          ON u.table_schema = s.table_schema
         AND u.table_name = s.table_name
         AND u.index_name = s.index_name
        WHERE s.table_schema = DATABASE()
        ORDER BY s.table_name, s.column_name;
        """
    )

    check_rows = mysql_exec(
        """
        SELECT tc.table_name, cc.check_clause
        FROM information_schema.table_constraints tc
        JOIN information_schema.check_constraints cc
          ON cc.constraint_schema = tc.constraint_schema
         AND cc.constraint_name = tc.constraint_name
        WHERE tc.constraint_schema = DATABASE()
          AND tc.constraint_type = 'CHECK'
        ORDER BY tc.table_name, tc.constraint_name;
        """
    )

    show_create = {table: get_show_create(table) for table in tables}

    return {
        "tables": tables,
        "table_comments": table_comments,
        "columns": column_rows,
        "foreign_keys": {(row[0], row[1]): f"{row[2]}.{row[3]}" for row in foreign_key_rows},
        "single_unique": {(row[0], row[1]) for row in unique_rows},
        "checks": defaultdict(list, {table: [] for table in tables}) | _group_checks(check_rows),
        "show_create": show_create,
    }


def _group_checks(rows: list[list[str]]) -> defaultdict[str, list[str]]:
    grouped: defaultdict[str, list[str]] = defaultdict(list)
    for table_name, clause in rows:
        grouped[table_name].append(clause)
    return grouped


def default_to_csv(default_value: str | None, extra: str) -> str:
    if default_value is None or default_value == "NULL":
        return ""
    if "CURRENT_TIMESTAMP" in default_value.upper():
        return default_value
    return default_value


def check_for_column(table_checks: list[str], column_name: str) -> str:
    hits = [clause for clause in table_checks if f"`{column_name}`" in clause]
    return " AND ".join(hits)


def build_design_rows(metadata: dict[str, Any]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for table_name, column_name, column_type, is_nullable, column_default, column_key, extra, column_comment, _ in metadata["columns"]:
        rows.append(
            {
                "数据表": table_name,
                "字段名": column_name,
                "数据类型": column_type.upper(),
                "是否有获取": "",
                "说明": column_comment,
                "PK": "1" if column_key == "PRI" else "",
                "NOT NULL": "1" if is_nullable == "NO" else "",
                "UNIQUE": "1" if (table_name, column_name) in metadata["single_unique"] else "",
                "DEFAULT": default_to_csv(column_default, extra),
                "FK": metadata["foreign_keys"].get((table_name, column_name), ""),
                "CHECK": check_for_column(metadata["checks"][table_name], column_name),
                "备注": "AUTO_INCREMENT" if "auto_increment" in extra.lower() else "",
                "父记录": metadata["table_comments"].get(table_name, ""),
            }
        )
    return rows


def rows_to_table_map(rows: list[dict[str, str]]) -> dict[str, list[dict[str, str]]]:
    table_map: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        table_map[row["数据表"]].append(row)
    return table_map


def insert_row_after(rows: list[dict[str, str]], field_name: str, new_row: dict[str, str]) -> None:
    for idx, row in enumerate(rows):
        if row["字段名"] == field_name:
            rows.insert(idx + 1, new_row)
            return
    rows.append(new_row)


def make_row(table_name: str, field_name: str, data_type: str, description: str, **kwargs: str) -> dict[str, str]:
    row = {
        "数据表": table_name,
        "字段名": field_name,
        "数据类型": data_type,
        "是否有获取": "",
        "说明": description,
        "PK": "",
        "NOT NULL": "",
        "UNIQUE": "",
        "DEFAULT": "",
        "FK": "",
        "CHECK": "",
        "备注": "",
        "父记录": "",
    }
    key_aliases = {
        "NOT_NULL": "NOT NULL",
    }
    for key, value in kwargs.items():
        row[key_aliases.get(key, key)] = value
    return row


def build_v2_design_rows(v1_rows: list[dict[str, str]], table_comments: dict[str, str]) -> list[dict[str, str]]:
    table_map = rows_to_table_map(deepcopy(v1_rows))

    for row in table_map["chain_industry"]:
        row["父记录"] = table_comments["chain_industry"]
    insert_row_after(
        table_map["chain_industry"],
        "chain_name",
        make_row(
            "chain_industry",
            "stage_key",
            "VARCHAR(32)",
            "产业阶段键（upstream/midstream/downstream）",
            NOT_NULL="1",
            DEFAULT="upstream",
            父记录=table_comments["chain_industry"],
        ),
    )
    insert_row_after(
        table_map["chain_industry"],
        "stage_key",
        make_row(
            "chain_industry",
            "stage_title",
            "VARCHAR(255)",
            "产业阶段标题",
            父记录=table_comments["chain_industry"],
        ),
    )
    insert_row_after(
        table_map["chain_industry"],
        "stage_title",
        make_row(
            "chain_industry",
            "sort_order",
            "INT",
            "排序序号",
            NOT_NULL="1",
            DEFAULT="0",
            CHECK="(`sort_order` >= 0)",
            父记录=table_comments["chain_industry"],
        ),
    )

    branch_rows = table_map["company_branch"]
    for row in branch_rows:
        if row["字段名"] == "company_branch_status":
            row["数据类型"] = "VARCHAR(64)"
            row["说明"] = "企业分支机构状态文本（如存续、注销、吊销）"
    insert_row_after(
        branch_rows,
        "company_branch_address",
        make_row(
            "company_branch",
            "company_branch_region",
            "VARCHAR(255)",
            "企业分支机构地区原始文本",
            父记录=table_comments["company_branch"],
        ),
    )

    customer_rows = table_map["company_customer"]
    insert_row_after(
        customer_rows,
        "company_customer_report_date",
        make_row(
            "company_customer",
            "data_source",
            "VARCHAR(255)",
            "数据来源",
            父记录=table_comments["company_customer"],
        ),
    )

    patent_rows = table_map["company_patent"]
    insert_row_after(
        patent_rows,
        "auth_date",
        make_row(
            "company_patent",
            "publication_date",
            "DATE",
            "专利申请公布日",
            父记录=table_comments["company_patent"],
        ),
    )

    qualification_rows = table_map["company_qualification"]
    insert_row_after(
        qualification_rows,
        "company_id",
        make_row(
            "company_qualification",
            "record_kind",
            "VARCHAR(32)",
            "记录类型（qualification/license）",
            NOT_NULL="1",
            DEFAULT="qualification",
            父记录=table_comments["company_qualification"],
        ),
    )
    insert_row_after(
        qualification_rows,
        "qualification_name",
        make_row(
            "company_qualification",
            "qualification_number",
            "VARCHAR(255)",
            "证书/许可编号",
            父记录=table_comments["company_qualification"],
        ),
    )
    insert_row_after(
        qualification_rows,
        "qualification_number",
        make_row(
            "company_qualification",
            "qualification_status",
            "VARCHAR(64)",
            "证书/许可状态",
            父记录=table_comments["company_qualification"],
        ),
    )
    insert_row_after(
        qualification_rows,
        "qualification_type",
        make_row(
            "company_qualification",
            "data_source",
            "VARCHAR(255)",
            "数据来源",
            父记录=table_comments["company_qualification"],
        ),
    )
    insert_row_after(
        qualification_rows,
        "issued_at",
        make_row(
            "company_qualification",
            "valid_from",
            "DATE",
            "有效期开始日期",
            父记录=table_comments["company_qualification"],
        ),
    )
    insert_row_after(
        qualification_rows,
        "expires_at",
        make_row(
            "company_qualification",
            "validity_period_text",
            "VARCHAR(255)",
            "有效期原始文本",
            父记录=table_comments["company_qualification"],
        ),
    )
    insert_row_after(
        qualification_rows,
        "validity_period_text",
        make_row(
            "company_qualification",
            "issuing_authority",
            "VARCHAR(255)",
            "发证/许可机关",
            父记录=table_comments["company_qualification"],
        ),
    )

    count_rows = table_map["company_basic_count"]
    insert_row_after(
        count_rows,
        "ranking_count_raw",
        make_row(
            "company_basic_count",
            "supplier_count",
            "INT",
            "供应商数量（结构化明细）",
            NOT_NULL="1",
            DEFAULT="0",
            父记录=table_comments["company_basic_count"],
        ),
    )
    insert_row_after(
        count_rows,
        "supplier_count",
        make_row(
            "company_basic_count",
            "supplier_count_raw",
            "INT",
            "供应商数量（原始聚合值）",
            NOT_NULL="1",
            DEFAULT="0",
            父记录=table_comments["company_basic_count"],
        ),
    )
    insert_row_after(
        count_rows,
        "supplier_count_raw",
        make_row(
            "company_basic_count",
            "qualification_count",
            "INT",
            "资质数量（结构化明细）",
            NOT_NULL="1",
            DEFAULT="0",
            父记录=table_comments["company_basic_count"],
        ),
    )
    insert_row_after(
        count_rows,
        "qualification_count",
        make_row(
            "company_basic_count",
            "qualification_count_raw",
            "INT",
            "资质数量（原始聚合值）",
            NOT_NULL="1",
            DEFAULT="0",
            父记录=table_comments["company_basic_count"],
        ),
    )
    insert_row_after(
        count_rows,
        "qualification_count_raw",
        make_row(
            "company_basic_count",
            "license_count",
            "INT",
            "许可数量（结构化明细）",
            NOT_NULL="1",
            DEFAULT="0",
            父记录=table_comments["company_basic_count"],
        ),
    )
    insert_row_after(
        count_rows,
        "license_count",
        make_row(
            "company_basic_count",
            "license_count_raw",
            "INT",
            "许可数量（原始聚合值）",
            NOT_NULL="1",
            DEFAULT="0",
            父记录=table_comments["company_basic_count"],
        ),
    )

    ordered_tables = list(dict.fromkeys(row["数据表"] for row in v1_rows))
    flat_rows: list[dict[str, str]] = []
    for table_name in ordered_tables:
        flat_rows.extend(table_map[table_name])
    return flat_rows


def _replace_line(sql: str, contains: str, new_line: str) -> str:
    lines = sql.splitlines()
    return "\n".join(new_line if contains in line else line for line in lines)


def _insert_after(sql: str, contains: str, new_line: str) -> str:
    lines = sql.splitlines()
    result: list[str] = []
    inserted = False
    for line in lines:
        result.append(line)
        if not inserted and contains in line:
            result.append(new_line)
            inserted = True
    if not inserted:
        raise ValueError(f"anchor not found: {contains}")
    return "\n".join(result)


def build_v2_create_sql(show_create: dict[str, str]) -> dict[str, str]:
    ddl = {table: sql for table, sql in show_create.items()}

    ddl["chain_industry"] = _insert_after(
        ddl["chain_industry"],
        "`chain_name` varchar(255)",
        "  `stage_key` varchar(32) COLLATE utf8mb4_unicode_ci NOT NULL DEFAULT 'upstream' COMMENT '产业阶段键（upstream/midstream/downstream）',",
    )
    ddl["chain_industry"] = _insert_after(
        ddl["chain_industry"],
        "`stage_key` varchar(32)",
        "  `stage_title` varchar(255) COLLATE utf8mb4_unicode_ci DEFAULT NULL COMMENT '产业阶段标题',",
    )
    ddl["chain_industry"] = _insert_after(
        ddl["chain_industry"],
        "`stage_title` varchar(255)",
        "  `sort_order` int NOT NULL DEFAULT '0' COMMENT '排序序号',",
    )

    ddl["company_branch"] = _insert_after(
        ddl["company_branch"],
        "`company_branch_address` varchar(255)",
        "  `company_branch_region` varchar(255) COLLATE utf8mb4_unicode_ci DEFAULT NULL COMMENT '企业分支机构地区原始文本',",
    )
    ddl["company_branch"] = _replace_line(
        ddl["company_branch"],
        "`company_branch_status` tinyint",
        "  `company_branch_status` varchar(64) COLLATE utf8mb4_unicode_ci DEFAULT NULL COMMENT '企业分支机构状态文本（如存续、注销、吊销）',",
    )

    ddl["company_customer"] = _insert_after(
        ddl["company_customer"],
        "`company_customer_report_date` date",
        "  `data_source` varchar(255) COLLATE utf8mb4_unicode_ci DEFAULT NULL COMMENT '数据来源',",
    )

    ddl["company_patent"] = _insert_after(
        ddl["company_patent"],
        "`auth_date` date",
        "  `publication_date` date DEFAULT NULL COMMENT '专利申请公布日',",
    )

    ddl["company_qualification"] = _insert_after(
        ddl["company_qualification"],
        "`company_id` bigint",
        "  `record_kind` varchar(32) COLLATE utf8mb4_unicode_ci NOT NULL DEFAULT 'qualification' COMMENT '记录类型（qualification/license）',",
    )
    ddl["company_qualification"] = _insert_after(
        ddl["company_qualification"],
        "`qualification_name` varchar(255)",
        "  `qualification_number` varchar(255) COLLATE utf8mb4_unicode_ci DEFAULT NULL COMMENT '证书/许可编号',",
    )
    ddl["company_qualification"] = _insert_after(
        ddl["company_qualification"],
        "`qualification_number` varchar(255)",
        "  `qualification_status` varchar(64) COLLATE utf8mb4_unicode_ci DEFAULT NULL COMMENT '证书/许可状态',",
    )
    ddl["company_qualification"] = _insert_after(
        ddl["company_qualification"],
        "`qualification_type` varchar(255)",
        "  `data_source` varchar(255) COLLATE utf8mb4_unicode_ci DEFAULT NULL COMMENT '数据来源',",
    )
    ddl["company_qualification"] = _insert_after(
        ddl["company_qualification"],
        "`issued_at` date",
        "  `valid_from` date DEFAULT NULL COMMENT '有效期开始日期',",
    )
    ddl["company_qualification"] = _insert_after(
        ddl["company_qualification"],
        "`expires_at` date",
        "  `validity_period_text` varchar(255) COLLATE utf8mb4_unicode_ci DEFAULT NULL COMMENT '有效期原始文本',",
    )
    ddl["company_qualification"] = _insert_after(
        ddl["company_qualification"],
        "`validity_period_text` varchar(255)",
        "  `issuing_authority` varchar(255) COLLATE utf8mb4_unicode_ci DEFAULT NULL COMMENT '发证/许可机关',",
    )
    ddl["company_qualification"] = _insert_after(
        ddl["company_qualification"],
        "PRIMARY KEY (`company_qualification_id`),",
        "  KEY `idx_company_qualification_company_kind` (`company_id`,`record_kind`),",
    )

    ddl["company_basic_count"] = _insert_after(
        ddl["company_basic_count"],
        "`ranking_count_raw` int",
        "  `supplier_count` int NOT NULL DEFAULT '0' COMMENT '供应商数量（结构化明细）',",
    )
    ddl["company_basic_count"] = _insert_after(
        ddl["company_basic_count"],
        "`supplier_count` int",
        "  `supplier_count_raw` int NOT NULL DEFAULT '0' COMMENT '供应商数量（原始聚合值）',",
    )
    ddl["company_basic_count"] = _insert_after(
        ddl["company_basic_count"],
        "`supplier_count_raw` int",
        "  `qualification_count` int NOT NULL DEFAULT '0' COMMENT '资质数量（结构化明细）',",
    )
    ddl["company_basic_count"] = _insert_after(
        ddl["company_basic_count"],
        "`qualification_count` int",
        "  `qualification_count_raw` int NOT NULL DEFAULT '0' COMMENT '资质数量（原始聚合值）',",
    )
    ddl["company_basic_count"] = _insert_after(
        ddl["company_basic_count"],
        "`qualification_count_raw` int",
        "  `license_count` int NOT NULL DEFAULT '0' COMMENT '许可数量（结构化明细）',",
    )
    ddl["company_basic_count"] = _insert_after(
        ddl["company_basic_count"],
        "`license_count` int",
        "  `license_count_raw` int NOT NULL DEFAULT '0' COMMENT '许可数量（原始聚合值）',",
    )

    return ddl


def seed_rows_v1() -> dict[str, list[dict[str, Any]]]:
    chain_seed = load_json(DATA_DIR / "chain_industry_seed.json")
    category_rows = load_json(DATA_DIR / "category_industry.json")
    category_by_code = {row["category_level_code"]: row["category_id"] for row in category_rows}
    company_tag_dimension = load_json(DATA_DIR / "company_tag_dimension.json")
    company_tag_subdimension = load_json(DATA_DIR / "company_tag_subdimension.json")
    company_tag_library = load_json(DATA_DIR / "company_tag_library.json")
    company_tag_auto_rule = load_json(DATA_DIR / "company_tag_auto_rule.json")

    return {
        "category_industry": category_rows,
        "chain_industry": [
            {
                "chain_id": row["chain_id"],
                "chain_name": row["chain_name"],
                "chain_des": row["chain_des"],
            }
            for row in chain_seed
        ],
        "chain_industry_category_industry_map": [
            {
                "chain_industry_category_industry_map_id": idx,
                "chain_id": row["chain_id"],
                "category_id": category_by_code[row["category_level_code"]],
            }
            for idx, row in enumerate(chain_seed, start=1)
        ],
        "company_tag_dimension": company_tag_dimension,
        "company_tag_subdimension": company_tag_subdimension,
        "company_tag_library": company_tag_library,
        "company_tag_auto_rule": [
            {
                "company_tag_auto_rule_id": idx,
                "company_tag_id": row["company_tag_id"],
                "company_tag_auto_rule_type": row["company_tag_auto_rule_type"],
                "rule_definition": row["rule_definition"],
                "is_enabled": row["is_enabled"],
            }
            for idx, row in enumerate(company_tag_auto_rule, start=1)
        ],
    }


def seed_rows_v2() -> dict[str, list[dict[str, Any]]]:
    chain_seed = load_json(DATA_DIR / "chain_industry_seed.json")
    category_rows = load_json(DATA_DIR / "category_industry.json")
    category_by_code = {row["category_level_code"]: row["category_id"] for row in category_rows}
    company_tag_dimension = load_json(DATA_DIR / "company_tag_dimension.json")
    company_tag_subdimension = load_json(DATA_DIR / "company_tag_subdimension.json")
    company_tag_library = load_json(DATA_DIR / "company_tag_library.json")
    company_tag_auto_rule = load_json(DATA_DIR / "company_tag_auto_rule.json")

    return {
        "category_industry": category_rows,
        "chain_industry": [
            {
                "chain_id": row["chain_id"],
                "chain_name": row["chain_name"],
                "stage_key": row["stage_key"],
                "stage_title": row["stage_title"],
                "chain_des": row["chain_des"],
                "sort_order": row["sort_order"],
            }
            for row in chain_seed
        ],
        "chain_industry_category_industry_map": [
            {
                "chain_industry_category_industry_map_id": idx,
                "chain_id": row["chain_id"],
                "category_id": category_by_code[row["category_level_code"]],
            }
            for idx, row in enumerate(chain_seed, start=1)
        ],
        "company_tag_dimension": company_tag_dimension,
        "company_tag_subdimension": company_tag_subdimension,
        "company_tag_library": company_tag_library,
        "company_tag_auto_rule": [
            {
                "company_tag_auto_rule_id": idx,
                "company_tag_id": row["company_tag_id"],
                "company_tag_auto_rule_type": row["company_tag_auto_rule_type"],
                "rule_definition": row["rule_definition"],
                "is_enabled": row["is_enabled"],
            }
            for idx, row in enumerate(company_tag_auto_rule, start=1)
        ],
    }


def build_init_sql(table_order: list[str], create_sql_by_table: dict[str, str], seed_rows: dict[str, list[dict[str, Any]]], seed_tables: list[str]) -> str:
    parts = [
        "/*!40101 SET @OLD_CHARACTER_SET_CLIENT=@@CHARACTER_SET_CLIENT */;",
        "/*!40101 SET @OLD_CHARACTER_SET_RESULTS=@@CHARACTER_SET_RESULTS */;",
        "/*!40101 SET @OLD_COLLATION_CONNECTION=@@COLLATION_CONNECTION */;",
        "/*!50503 SET NAMES utf8mb4 */;",
        "/*!40103 SET @OLD_TIME_ZONE=@@TIME_ZONE */;",
        "/*!40103 SET TIME_ZONE='+00:00' */;",
        "/*!40014 SET @OLD_UNIQUE_CHECKS=@@UNIQUE_CHECKS, UNIQUE_CHECKS=0 */;",
        "/*!40014 SET @OLD_FOREIGN_KEY_CHECKS=@@FOREIGN_KEY_CHECKS, FOREIGN_KEY_CHECKS=0 */;",
        "/*!40101 SET @OLD_SQL_MODE=@@SQL_MODE, SQL_MODE='NO_AUTO_VALUE_ON_ZERO' */;",
        "/*!40111 SET @OLD_SQL_NOTES=@@SQL_NOTES, SQL_NOTES=0 */;",
        "",
        "CREATE DATABASE IF NOT EXISTS `industrial_chain` DEFAULT CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci;",
        "USE `industrial_chain`;",
        "",
    ]

    for table_name in table_order:
        parts.append(f"DROP TABLE IF EXISTS `{table_name}`;")
        parts.append(create_sql_by_table[table_name] + ";")
        parts.append("")

    for table_name in seed_tables:
        if table_name in seed_rows and seed_rows[table_name]:
            parts.append(insert_statement(table_name, seed_rows[table_name]))

    parts.extend(
        [
            "/*!40103 SET TIME_ZONE=@OLD_TIME_ZONE */;",
            "/*!40101 SET SQL_MODE=@OLD_SQL_MODE */;",
            "/*!40014 SET FOREIGN_KEY_CHECKS=@OLD_FOREIGN_KEY_CHECKS */;",
            "/*!40014 SET UNIQUE_CHECKS=@OLD_UNIQUE_CHECKS */;",
            "/*!40101 SET CHARACTER_SET_CLIENT=@OLD_CHARACTER_SET_CLIENT */;",
            "/*!40101 SET CHARACTER_SET_RESULTS=@OLD_CHARACTER_SET_RESULTS */;",
            "/*!40101 SET COLLATION_CONNECTION=@OLD_COLLATION_CONNECTION */;",
            "/*!40111 SET SQL_NOTES=@OLD_SQL_NOTES */;",
            "",
        ]
    )
    return "\n".join(parts)


def write_design_csv(path: Path, rows: list[dict[str, str]]) -> None:
    fieldnames = ["数据表", "字段名", "数据类型", "是否有获取", "说明", "PK", "NOT NULL", "UNIQUE", "DEFAULT", "FK", "CHECK", "备注", "父记录"]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def collect_xlsx_summary() -> dict[str, Any]:
    workbook = load_workbook(UNCLEAN_XLSX, read_only=True, data_only=True)
    sheet_rows: dict[str, int] = {}
    sheet_headers: dict[str, list[str]] = {}
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        sheet_rows[sheet_name] = ws.max_row - 1
        sheet_headers[sheet_name] = [str(value) if value is not None else "" for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    return {"rows": sheet_rows, "headers": sheet_headers}


def write_report(v1_design_rows: list[dict[str, str]], v2_design_rows: list[dict[str, str]], tag_counts: dict[str, int], xlsx_summary: dict[str, Any]) -> None:
    v1_tables = {row["数据表"] for row in v1_design_rows}
    v2_tables = {row["数据表"] for row in v2_design_rows}
    v1_by_table = rows_to_table_map(v1_design_rows)
    v2_by_table = rows_to_table_map(v2_design_rows)
    added_tables = sorted(v2_tables - v1_tables)
    changed_tables = []
    for table_name in sorted(v1_tables & v2_tables):
        v1_columns = [row["字段名"] for row in v1_by_table[table_name]]
        v2_columns = [row["字段名"] for row in v2_by_table[table_name]]
        if v1_columns != v2_columns or any(a != b for a, b in zip(v1_by_table[table_name], v2_by_table[table_name])):
            changed_tables.append(table_name)

    content = f"""# 设计变更说明（2026-04-05）

## 1. 基线校对结论

- 旧版 `SQL/sql/init.sql` 不是干净的初始化脚本，内部存在重复 `CREATE TABLE` 片段，已不适合作为可信基线。
- 活库比 `SQL/mysql-design/design-2026-4-1-V1.csv` 多出完整的 `platform_notice` 表结构，因此 `design-2026-4-5-V1.csv` 已按本地 MySQL 重新校准。
- `SQL/data` 中的企业标签静态源数据落后于活库，已补齐 1 个维度、2 个子维度、51 个标签项，避免后续重建数据库时丢失“行业标签”能力。
- `chain_industry_seed.json` 含 `stage_key` / `stage_title` / `sort_order`，旧版 `chain_industry` 无法完整表达该源数据。

## 2. 最新 Excel 的主要结构冲突

- `分支机构`：最新数据提供文本状态和地区原始值，旧表 `company_branch.company_branch_status` 为 `TINYINT`，且没有地区字段。
- `许可`：最新数据是独立工作表，旧库没有单独承载许可元数据的字段集。
- `资质`：最新数据包含 `证书编号` 与 `证书状态`，旧表 `company_qualification` 无对应字段。
- `企业经营信息`：最新数据提供 `供应商数量`、`许可数量`、`资质数量`，旧表 `company_basic_count` 没有对应聚合字段。
- `客户信息`：最新数据提供 `数据来源`，旧表 `company_customer` 无对应字段。
- `专利信息`：最新数据提供 `申请公布日`，旧表 `company_patent` 只有 `auth_date`（专利授权日期），语义不匹配。

## 3. V2 设计调整

- `chain_industry`：新增 `stage_key`、`stage_title`、`sort_order`，使表结构与 `SQL/data/chain_industry_seed.json` 对齐。
- `company_branch`：将 `company_branch_status` 改为文本类型，并新增 `company_branch_region`。
- `company_qualification`：不新建 `company_license` 表，改为扩展现有证照表，新增 `record_kind`、`qualification_number`、`qualification_status`、`data_source`、`valid_from`、`validity_period_text`、`issuing_authority`，统一承载“许可/资质”两类记录。
- `company_basic_count`：新增 `supplier_count` / `supplier_count_raw`、`qualification_count` / `qualification_count_raw`、`license_count` / `license_count_raw`。
- `company_customer`：新增 `data_source`。
- `company_patent`：新增 `publication_date`，避免把“申请公布日”误写进“授权日期”。

## 4. 本次同步的静态源数据

- 企业标签维度总数：{tag_counts["dimensions"]}
- 企业标签子维度总数：{tag_counts["subdimensions"]}
- 企业标签库总数：{tag_counts["library"]}
- 行业分类源：`SQL/data/category_industry.json`
- 产业链源：`SQL/data/chain_industry_seed.json`
- 企业标签源：`SQL/data/company_tag_dimension.json`、`SQL/data/company_tag_subdimension.json`、`SQL/data/company_tag_library.json`、`SQL/data/company_tag_auto_rule.json`

## 5. Excel 工作表概览

"""
    for sheet_name, row_count in xlsx_summary["rows"].items():
        headers = "、".join(xlsx_summary["headers"][sheet_name])
        content += f"- `{sheet_name}`：{row_count} 行；字段：{headers}\n"

    content += "\n## 6. 产物清单\n\n"
    content += f"- `SQL/mysql-design/{DESIGN_V1_PATH.name}`：基于旧 MySQL 的校准设计\n"
    content += f"- `SQL/mysql-design/{DESIGN_V2_PATH.name}`：支持最新 Excel 与静态源数据的目标设计\n"
    content += f"- `SQL/sql/{INIT_V1_PATH.name}`：基于旧 MySQL 结构整理出的初始化脚本快照\n"
    content += f"- `SQL/sql/{INIT_V2_PATH.name}`：V2 目标初始化脚本（未导入最新企业数据）\n"
    content += f"- `SQL/reports/{REPORT_PATH.name}`：本说明文件\n"
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text(content, encoding="utf-8")


def main() -> None:
    DESIGN_DIR.mkdir(parents=True, exist_ok=True)
    SQL_OUT_DIR.mkdir(parents=True, exist_ok=True)
    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    tag_counts = sync_tag_source_files()
    metadata = live_schema_metadata()

    v1_design_rows = build_design_rows(metadata)
    write_design_csv(DESIGN_V1_PATH, v1_design_rows)

    v2_design_rows = build_v2_design_rows(v1_design_rows, metadata["table_comments"])
    write_design_csv(DESIGN_V2_PATH, v2_design_rows)

    v1_init_sql = build_init_sql(metadata["tables"], metadata["show_create"], seed_rows_v1(), SEED_TABLES_V1)
    INIT_V1_PATH.write_text(v1_init_sql, encoding="utf-8")

    v2_create_sql = build_v2_create_sql(metadata["show_create"])
    v2_init_sql = build_init_sql(metadata["tables"], v2_create_sql, seed_rows_v2(), SEED_TABLES_V2)
    INIT_V2_PATH.write_text(v2_init_sql, encoding="utf-8")

    write_report(v1_design_rows, v2_design_rows, tag_counts, collect_xlsx_summary())


if __name__ == "__main__":
    main()
